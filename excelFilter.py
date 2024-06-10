import pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def filter_and_copy_excel(file_path, sheet_name, column_name, filter_value, output_sheet_name):
    try:
        # Load the Excel file
        xls = pd.ExcelFile(file_path)

        # Check if the specified sheet exists
        if sheet_name not in xls.sheet_names:
            messagebox.showerror("Error", f"Sheet '{sheet_name}' does not exist in the Excel file.")
            return

        # Read the specified sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Check if the specified column exists
        if column_name not in df.columns:
            messagebox.showerror("Error", f"Column '{column_name}' does not exist in the sheet '{sheet_name}'.")
            return

        # Add "Filtered" column if it does not exist
        if 'Filtered' not in df.columns:
            df['Filtered'] = None

        # Filter the DataFrame based on the specified column and value (partial match) and not previously filtered
        filtered_df = df[df[column_name].str.contains(filter_value, case=False, na=False)]

        # Check if any rows are filtered
        num_rows_copied = len(filtered_df)

        if num_rows_copied == 0:
            messagebox.showinfo("No Results",
                                f"No new rows found containing '{filter_value}' in column '{column_name}'.")
            return

        # Mark filtered rows with the output sheet name in the original DataFrame
        def append_sheet_name(existing_value):
            if pd.isna(existing_value):
                return output_sheet_name
            else:
                existing_sheets = existing_value.split(", ")
                if output_sheet_name not in existing_sheets:
                    existing_sheets.append(output_sheet_name)
                return ", ".join(existing_sheets)

        df.loc[filtered_df.index, 'Filtered'] = df.loc[filtered_df.index, 'Filtered'].apply(append_sheet_name)

        # Remove the "Filtered" column from the filtered DataFrame before copying
        filtered_df = filtered_df.drop(columns=['Filtered'])

        # Create a Pandas Excel writer using the same file path
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Write the filtered DataFrame to a new sheet in the same Excel file
            filtered_df.to_excel(writer, sheet_name=output_sheet_name, index=False)
            # Write the updated original DataFrame back to its original sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        messagebox.showinfo("Success",
                            f"Filtered data written to sheet '{output_sheet_name}' in the Excel file.\n{num_rows_copied} rows copied.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def auto_detect_and_copy(file_path, sheet_name, column_name):
    try:
        # Load the Excel file
        xls = pd.ExcelFile(file_path)

        # Check if the specified sheet exists
        if sheet_name not in xls.sheet_names:
            messagebox.showerror("Error", f"Sheet '{sheet_name}' does not exist in the Excel file.")
            return

        # Read the specified sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Check if the specified column exists
        if column_name not in df.columns:
            messagebox.showerror("Error", f"Column '{column_name}' does not exist in the sheet '{sheet_name}'.")
            return

        # Add "Filtered" column if it does not exist
        if 'Filtered' not in df.columns:
            df['Filtered'] = None

        # Process each sheet in the Excel file except the input sheet
        for output_sheet_name in xls.sheet_names:
            if output_sheet_name == sheet_name:
                continue

            # Filter the DataFrame based on the column and sheet name (partial match)
            filtered_df = df[df[column_name].str.contains(output_sheet_name, case=False, na=False)]

            # Check if any rows are filtered
            num_rows_copied = len(filtered_df)

            if num_rows_copied == 0:
                continue

            # Mark filtered rows with the output sheet name in the original DataFrame
            def append_sheet_name(existing_value):
                if pd.isna(existing_value):
                    return output_sheet_name
                else:
                    existing_sheets = existing_value.split(", ")
                    if output_sheet_name not in existing_sheets:
                        existing_sheets.append(output_sheet_name)
                    return ", ".join(existing_sheets)

            df.loc[filtered_df.index, 'Filtered'] = df.loc[filtered_df.index, 'Filtered'].apply(append_sheet_name)

            # Remove the "Filtered" column from the filtered DataFrame before copying
            filtered_df = filtered_df.drop(columns=['Filtered'])

            # Create a Pandas Excel writer using the same file path
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Write the filtered DataFrame to a new sheet in the same Excel file
                filtered_df.to_excel(writer, sheet_name=output_sheet_name, index=False)
                # Write the updated original DataFrame back to its original sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Load the workbook to apply styles
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]

        # Define a red fill for cells
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        # Apply the red fill to rows with no or more than one entry in the "Filtered" column
        for row in range(2, worksheet.max_row + 1):
            filtered_cell = worksheet.cell(row=row, column=worksheet.max_column)
            filtered_value = filtered_cell.value
            if pd.isna(filtered_value) or len(filtered_value.split(", ")) != 1:
                for col in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row, column=col).fill = red_fill

        workbook.save(file_path)

        messagebox.showinfo("Success",
                            "Auto-detection and copying completed successfully.\nRows with no or more than one entry in the 'Filtered' column have been marked.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def open_file_dialog_manual():
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    if file_path:
        try:
            # Load the Excel file
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names
            sheet_name = select_option("Select Sheet", "Select the sheet to process:", sheet_names)
            if not sheet_name:
                return

            # Read the specified sheet into a DataFrame
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            column_names = df.columns.tolist()
            column_name = select_option("Select Column", "Select the column to filter by:", column_names)
            if not column_name:
                return

            # Get the filter value and output sheet name from the user
            filter_value = simpledialog.askstring("Input", "Enter the value to filter for:")
            output_sheet_name = simpledialog.askstring("Input", "Enter the output sheet name:")
            if filter_value and output_sheet_name:
                summary = (
                    f"File Path: {file_path}\n"
                    f"Sheet Name: {sheet_name}\n"
                    f"Column Name: {column_name}\n"
                    f"Filter Value: {filter_value}\n"
                    f"Output Sheet Name: {output_sheet_name}"
                )
                if messagebox.askokcancel("Confirm Details", f"Please confirm the details:\n\n{summary}"):
                    filter_and_copy_excel(file_path, sheet_name, column_name, filter_value, output_sheet_name)
            else:
                messagebox.showerror("Error", "Filter value and output sheet name must be provided.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")


def open_file_dialog_auto():
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    if file_path:
        try:
            # Load the Excel file
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names
            sheet_name = select_option("Select Sheet", "Select the input sheet:", sheet_names)
            if not sheet_name:
                return

            # Read the specified sheet into a DataFrame
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            column_names = df.columns.tolist()
            column_name = select_option("Select Column", "Select the column to filter by:", column_names)
            if not column_name:
                return

            summary = (
                f"File Path: {file_path}\n"
                f"Sheet Name: {sheet_name}\n"
                f"Column Name: {column_name}"
            )
            if messagebox.askokcancel("Confirm Details", f"Please confirm the details:\n\n{summary}"):
                auto_detect_and_copy(file_path, sheet_name, column_name)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")


def select_option(title, prompt, options):
    option_window = tk.Toplevel(root)
    option_window.title(title)
    option_window.geometry("300x200")
    option_window.configure(bg="#f0f0f0")

    tk.Label(option_window, text=prompt, bg="#f0f0f0").pack(pady=10)

    selected_option = tk.StringVar(option_window)
    selected_option.set(options[0])

    option_menu = ttk.OptionMenu(option_window, selected_option, options[0], *options)
    option_menu.pack(pady=10)

    tk.Button(option_window, text="Confirm", command=option_window.destroy).pack(pady=10)

    option_window.wait_window()
    return selected_option.get()


def show_readme():
    readme_text = (
        "Excel Filter and Copy Tool\n\n"
        "Manual Mode:\n"
        "1. Click 'Select Excel File (Manual)'.\n"
        "2. Choose the Excel file to process.\n"
        "3. Select the sheet containing the data from the dropdown.\n"
        "4. Select the column to filter by from the dropdown.\n"
        "5. Enter the value to filter for.\n"
        "6. Enter the output sheet name to save the filtered data.\n\n"
        "Auto Detect Mode:\n"
        "1. Click 'Select Excel File (Auto Detect)'.\n"
        "2. Choose the Excel file to process.\n"
        "3. Select the sheet containing the data from the dropdown.\n"
        "4. Select the column to filter by from the dropdown.\n"
        "5. The program will automatically filter and copy data to sheets named after the detected values.\n"
        "6. Rows with no or more than one entry in the 'Filtered' column will be marked with a red background.\n\n"
        "Note: The program will update the original sheet by marking the filtered rows."
    )
    messagebox.showinfo("Readme", readme_text)


# Create the main window
root = tk.Tk()
root.title("Excel Filter and Copy")
root.geometry("300x200")
root.configure(bg="#f0f0f0")

style = ttk.Style()
style.configure("TButton", padding=6, relief="flat", background="#ccc")

# Add buttons to open the file dialog for manual and auto modes
open_button_manual = ttk.Button(root, text="Select Excel File (Manual)", command=open_file_dialog_manual)
open_button_manual.pack(pady=10)

open_button_auto = ttk.Button(root, text="Select Excel File (Auto Detect)", command=open_file_dialog_auto)
open_button_auto.pack(pady=10)

readme_button = ttk.Button(root, text="Readme", command=show_readme)
readme_button.pack(pady=10)

# Run the Tkinter event loop
root.mainloop()
